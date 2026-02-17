# ============================================
# Archivo: modulos/importaciones_excel.py
# Propósito:
#   - Importar clientes desde Excel (.xlsx) usando openpyxl
#   - Crear o actualizar según rut_normalizado
# ============================================

from __future__ import annotations  # Tipos modernos
from typing import Any, Dict, List  # Tipos para claridad

from openpyxl import load_workbook  # Leer archivos Excel

from .repo_clientes import (
    crear_cliente,                               # Inserta cliente nuevo
    actualizar_cliente,                           # Actualiza cliente existente
    obtener_cliente_id_por_rut_normalizado,       # Detecta si ya existe por RUT
)
from .validaciones import normalizar_texto, rut_a_normalizado  # Limpieza y normalización RUT
from .repo_logs import registrar_evento  # Registrar eventos en logs


def _leer_filas_excel(ruta_xlsx: str) -> List[Dict[str, Any]]:
    """
    Lee un Excel y devuelve una lista de diccionarios (uno por fila).
    - La primera fila se toma como encabezados.
    - Los encabezados se normalizan a minúsculas para compararlos fácil.

    Encabezados recomendados:
      tipo_cliente, rut, nombres, apellidos, razon_social,
      email, telefono, estado, recibe_correos, observaciones
    """
    wb = load_workbook(ruta_xlsx)   # Abre el archivo Excel
    ws = wb.active                 # Usa la primera hoja

    # Leemos encabezados de la fila 1
    headers: List[str] = []
    for cell in ws[1]:
        headers.append(normalizar_texto(cell.value).casefold())

    registros: List[Dict[str, Any]] = []

    # Recorremos filas desde la 2 (datos)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Si la fila está vacía, la saltamos
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        registro: Dict[str, Any] = {}
        for h, v in zip(headers, row):
            registro[h] = v

        registros.append(registro)

    return registros


def importar_clientes_excel(ruta_xlsx: str, modo: str = "actualizar") -> Dict[str, Any]:
    """
    FUNCIÓN PRINCIPAL (la que importa la UI)

    Importa clientes desde Excel.

    modo:
      - "actualizar": si existe el RUT -> actualiza; si no existe -> crea.
      - "rechazar":   si existe el RUT -> rechaza; si no existe -> crea.

    Retorna resumen:
      {"agregados": int, "actualizados": int, "rechazados": int, "errores": [str, ...]}
    """
    registrar_evento("importacion", "INICIO", f"Importación Excel: {ruta_xlsx}")

    filas = _leer_filas_excel(ruta_xlsx)

    agregados = 0
    actualizados = 0
    rechazados = 0
    errores: List[str] = []

    # idx inicia en 2 porque la fila 1 es encabezado
    for idx, fila in enumerate(filas, start=2):
        try:
            # Armamos el dict con las claves que espera el CRUD/validación
            datos = {
                "tipo_cliente": fila.get("tipo_cliente"),
                "rut": fila.get("rut"),
                "nombres": fila.get("nombres"),
                "apellidos": fila.get("apellidos"),
                "razon_social": fila.get("razon_social"),
                "email": fila.get("email"),
                "telefono": fila.get("telefono"),
                "estado": fila.get("estado", 1),
                "recibe_correos": fila.get("recibe_correos", 1),
                "observaciones": fila.get("observaciones", ""),
            }

            # Normalizamos RUT para detectar existencia
            rut_norm = rut_a_normalizado(normalizar_texto(datos.get("rut")))
            if not rut_norm:
                raise ValueError("RUT vacío.")

            # ¿Existe ya?
            existente_id = obtener_cliente_id_por_rut_normalizado(rut_norm)

            if existente_id is None:
                # No existe -> crear
                crear_cliente(datos)
                agregados += 1
            else:
                # Existe -> actualizar o rechazar
                if modo == "rechazar":
                    rechazados += 1
                else:
                    actualizar_cliente(existente_id, datos)
                    actualizados += 1

        except Exception as e:
            rechazados += 1
            errores.append(f"Fila {idx}: {e}")

    resumen = {
        "agregados": agregados,
        "actualizados": actualizados,
        "rechazados": rechazados,
        "errores": errores,
    }

    registrar_evento(
        "importacion",
        "FIN",
        f"Agregados={agregados} Actualizados={actualizados} Rechazados={rechazados}",
    )

    return resumen
